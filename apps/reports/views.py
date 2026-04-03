from datetime import date
from django.http import HttpResponse
from django.db.models import Count, Sum, Q
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apps.accounts.models import User, Department
from apps.accounts.permissions import IsAdminOrSupervisor
from apps.attendance.models import AttendanceRecord
from apps.leave.models import LeaveRequest


def parse_month_param(request):
    month_param = request.query_params.get('month')
    if month_param:
        try:
            year, month = month_param.split('-')
            return int(year), int(month)
        except ValueError:
            return None, None
    from django.utils import timezone
    now = timezone.now()
    return now.year, now.month


class AttendanceReportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrSupervisor]

    def get(self, request):
        year, month = parse_month_param(request)
        if not year:
            return Response({'detail': 'Format bulan tidak valid. Gunakan YYYY-MM.'}, status=400)

        dept_id = request.query_params.get('department')
        staff_id = request.query_params.get('staff_id')

        records = AttendanceRecord.objects.filter(
            date__year=year, date__month=month
        ).select_related('staff', 'staff__department', 'schedule')

        if dept_id:
            records = records.filter(staff__department=dept_id)
        if staff_id:
            records = records.filter(staff=staff_id)

        data = []
        for r in records:
            work_dur = ''
            if r.work_duration:
                total_sec = int(r.work_duration.total_seconds())
                h, rem = divmod(total_sec, 3600)
                m, _ = divmod(rem, 60)
                work_dur = f'{h}j {m}m'
            data.append({
                'id': r.id,
                'date': r.date,
                'employee_id': r.staff.employee_id,
                'staff_name': r.staff.full_name,
                'department': r.staff.department.name if r.staff.department else '-',
                'schedule': r.schedule.name if r.schedule else '-',
                'check_in': r.check_in_time.astimezone().strftime('%H:%M:%S') if r.check_in_time else '-',
                'check_out': r.check_out_time.astimezone().strftime('%H:%M:%S') if r.check_out_time else '-',
                'status': r.get_status_display(),
                'work_duration': work_dur,
                'is_overtime': r.is_overtime,
                'notes': r.notes or '',
            })

        return Response({
            'month': f'{year}-{month:02d}',
            'total_records': len(data),
            'records': data,
        })


class AttendanceReportExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrSupervisor]

    def get(self, request):
        year, month = parse_month_param(request)
        if not year:
            return Response({'detail': 'Format bulan tidak valid. Gunakan YYYY-MM.'}, status=400)

        dept_id = request.query_params.get('department')
        records = AttendanceRecord.objects.filter(
            date__year=year, date__month=month
        ).select_related('staff', 'staff__department', 'schedule').order_by('date', 'staff__full_name')

        if dept_id:
            records = records.filter(staff__department=dept_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Presensi {year}-{month:02d}'

        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            'No', 'Tanggal', 'NIP', 'Nama', 'Departemen', 'Jadwal',
            'Check-in', 'Check-out', 'Status', 'Durasi Kerja', 'Lembur', 'Catatan'
        ]
        col_widths = [5, 13, 15, 25, 20, 18, 10, 10, 15, 14, 8, 30]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        for row_idx, r in enumerate(records, 2):
            work_dur = ''
            if r.work_duration:
                total_sec = int(r.work_duration.total_seconds())
                h, rem = divmod(total_sec, 3600)
                m, _ = divmod(rem, 60)
                work_dur = f'{h}j {m}m'

            row_data = [
                row_idx - 1,
                r.date.strftime('%d/%m/%Y'),
                r.staff.employee_id,
                r.staff.full_name,
                r.staff.department.name if r.staff.department else '-',
                r.schedule.name if r.schedule else '-',
                r.check_in_time.astimezone().strftime('%H:%M:%S') if r.check_in_time else '-',
                r.check_out_time.astimezone().strftime('%H:%M:%S') if r.check_out_time else '-',
                r.get_status_display(),
                work_dur,
                'Ya' if r.is_overtime else 'Tidak',
                r.notes or '',
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx in (1, 7, 8, 10, 11):
                    cell.alignment = center

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:L{len(records) + 1}'

        month_name = [
            'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
            'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
        ][month - 1]

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Presensi_{month_name}_{year}.xlsx"'
        wb.save(response)
        return response


class LeaveSummaryReportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrSupervisor]

    def get(self, request):
        year = int(request.query_params.get('year', date.today().year))
        dept_id = request.query_params.get('department')

        leave_qs = LeaveRequest.objects.filter(
            start_date__year=year, status='disetujui'
        ).select_related('staff', 'leave_type', 'staff__department')

        if dept_id:
            leave_qs = leave_qs.filter(staff__department=dept_id)

        data = []
        for r in leave_qs:
            data.append({
                'id': r.id,
                'employee_id': r.staff.employee_id,
                'staff_name': r.staff.full_name,
                'department': r.staff.department.name if r.staff.department else '-',
                'leave_type': r.leave_type.name,
                'start_date': r.start_date,
                'end_date': r.end_date,
                'total_days': r.total_days,
                'reason': r.reason,
                'reviewed_by': r.reviewed_by.full_name if r.reviewed_by else '-',
            })

        return Response({
            'year': year,
            'total_requests': len(data),
            'records': data,
        })


class DepartmentSummaryReportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrSupervisor]

    def get(self, request):
        year, month = parse_month_param(request)
        if not year:
            return Response({'detail': 'Format bulan tidak valid. Gunakan YYYY-MM.'}, status=400)

        departments = Department.objects.all()
        result = []

        for dept in departments:
            base_qs = AttendanceRecord.objects.filter(
                staff__department=dept,
                date__year=year,
                date__month=month
            )
            counts = base_qs.aggregate(
                total=Count('id'),
                hadir=Count('id', filter=Q(status='hadir')),
                terlambat=Count('id', filter=Q(status='terlambat')),
                pulang_awal=Count('id', filter=Q(status='pulang_awal')),
                absen=Count('id', filter=Q(status='absen')),
                izin=Count('id', filter=Q(status='izin')),
                libur=Count('id', filter=Q(status='libur')),
                wfh=Count('id', filter=Q(status='wfh')),
            )
            result.append({
                'department_id': dept.id,
                'department_name': dept.name,
                'department_code': dept.code,
                'staff_count': dept.staff.filter(is_active=True).count(),
                **counts,
            })

        return Response({
            'month': f'{year}-{month:02d}',
            'departments': result,
        })
